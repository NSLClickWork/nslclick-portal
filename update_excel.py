import openpyxl

file_path = 'NSL_Bot_Tracker.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

for row in range(1, sheet.max_row + 1):
    bot_name = sheet.cell(row=row, column=2).value
    if bot_name == 'Recruiting Bot':
        sheet.cell(row=row, column=5).value = "• Candidate intake form\n• Status tracking\n• Notification on new applications\n• Syncs with Web Portal\n• AI CV Grading (Chấm điểm CV)"
        sheet.cell(row=row, column=6).value = "Web Portal (candidate DB)"
        sheet.cell(row=row, column=8).value = "Integrated into web portal with AI CV grading."

wb.save(file_path)
print("Excel file updated successfully.")
